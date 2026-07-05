"""Behavioral smoke test for v1.7「Office 体系 2.0」— the 字体纪律 rebuild of write_document / write_excel.

用户真机反馈:AI 产出的 docx「字体都不对，中文大小不一」。把关人取证:通篇零字体声明 → Word 各级默认回落。
v1.7 在 styles.xml 级注入显式字体 (Normal/Title/Heading 1-3 全部 rFonts ascii+hAnsi+eastAsia = 令牌
微软雅黑 + 字号阶梯 + 标题主色),让任何 run 不再裸奔。write_excel 落笔即样式 (token 字体 + 数字格式启发式)。

本冒烟直接回读 OOXML 断言这些结构事实:

  ① Word 字体纪律 (硬断言项):
     * styles.xml 中 Normal / Title / Heading 1-3 的 rFonts w:eastAsia == 微软雅黑 (中文不再回落宋体)。
     * 字号阶梯正确: Normal 11 / H3 14 / H2 16 / H1 20 / Title 28 pt (sz 以半点存: 22/28/32/40/56)。
     * 标题主色 (business 深蓝) 落在 Title / Heading 的 w:color。
     * 无裸 run: document.xml 里每个 <w:r> 要么自带 rFonts,要么其段落用的是已注入字体的内建样式 →
       抽查正文/标题/表格所有 run 的「有效 eastAsia 字体」都解析到 微软雅黑 (非空且非默认宋体)。
     * 封面深底满铺 + 标题下 pBdr 细横线 + 「第 X 页」PAGE 域页脚 都在。
  ② Excel 落笔即样式: 全表 (含表头与数据) Font.name == 微软雅黑;数字列拿到启发式数字格式
     (货币 / 百分比 / 千分位);数字型字符串被转成真数字。
  ③ 三套设计系统: render_samples 三风格 × 4 件 = 12 件全部生成且非空。
  ④ 向后兼容: 老式 write_document(title+content 纯文本/#heading) 与 write_excel(headers+data) 照常
     成功,且自动吃上新样式 (eastAsia 到位)。

Run with UTF-8:  python -X utf8 tests/smoke_v17.py
"""

import os
import re
import sys
import tempfile
import zipfile

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(_ROOT, "src")
sys.path.insert(0, _SRC)

_DATA = os.path.join(tempfile.gettempdir(), "acc_smoke_v17_data")
os.makedirs(_DATA, exist_ok=True)
os.environ["WCW_DATA_DIR"] = _DATA

import ai_computer_control.server as server  # noqa: E402
from ai_computer_control.tools import office_style as tokens  # noqa: E402

_FNS = {t.name: t.fn for t in server.mcp._tool_manager.list_tools()}
_FAILURES: list[str] = []

_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def check(cond: bool, msg: str):
    print(f"  [{'ok  ' if cond else 'FAIL'}] {msg}")
    if not cond:
        _FAILURES.append(msg)


# ---------------------------------------------------------------------------------------------
# Word OOXML helpers: parse styles.xml + document.xml with lxml and resolve each run's *effective*
# eastAsia font (direct rFonts on the run, else the paragraph's style's rFonts, else Normal's).
# ---------------------------------------------------------------------------------------------
def _style_ea_map(styles_root):
    """styleId -> eastAsia font (from that style's rPr/rFonts@w:eastAsia). Missing -> None."""
    out = {}
    for st in styles_root.findall(_W + "style"):
        sid = st.get(_W + "styleId")
        rpr = st.find(_W + "rPr")
        ea = None
        if rpr is not None:
            rf = rpr.find(_W + "rFonts")
            if rf is not None:
                ea = rf.get(_W + "eastAsia")
        out[sid] = ea
    return out


def _style_field(styles_root, style_id, field):
    """Return (eastAsia, sz-halfpoints, color) for one style, or (None, None, None)."""
    for st in styles_root.findall(_W + "style"):
        if st.get(_W + "styleId") != style_id:
            continue
        rpr = st.find(_W + "rPr")
        if rpr is None:
            return None
        if field == "ea":
            rf = rpr.find(_W + "rFonts")
            return rf.get(_W + "eastAsia") if rf is not None else None
        if field == "sz":
            sz = rpr.find(_W + "sz")
            return sz.get(_W + "val") if sz is not None else None
        if field == "color":
            col = rpr.find(_W + "color")
            return col.get(_W + "val") if col is not None else None
    return None


def _docx_font_discipline(path: str, expect_font: str, expect_title_color: str):
    """Assert the write_document 字体纪律 on a real docx: style-level eastAsia/sz/color + no 裸 run."""
    from lxml import etree

    with zipfile.ZipFile(path) as z:
        styles_root = etree.fromstring(z.read("word/styles.xml"))
        doc_root = etree.fromstring(z.read("word/document.xml"))
        parts = {n: z.read(n) for n in z.namelist()}

    # --- style-level eastAsia + size ladder + colour ---
    ladder = {"Normal": "22", "Heading3": "28", "Heading2": "32", "Heading1": "40", "Title": "56"}
    for sid, want_sz in ladder.items():
        ea = _style_field(styles_root, sid, "ea")
        sz = _style_field(styles_root, sid, "sz")
        check(ea == expect_font, f"style {sid}: eastAsia == {expect_font} (got {ea!r})")
        check(sz == want_sz, f"style {sid}: sz == {want_sz} half-pt (got {sz!r})")
    # heading / title colour is the business deep-blue family
    for sid in ("Title", "Heading1"):
        col = _style_field(styles_root, sid, "color")
        check(col is not None and col.upper() != "AUTO",
              f"style {sid}: explicit colour set (got {col!r})")

    # --- no 裸 run: every <w:r> resolves to a non-default eastAsia font ---
    ea_by_style = _style_ea_map(styles_root)
    normal_ea = ea_by_style.get("Normal")
    bare = 0
    total = 0
    body = doc_root.find(_W + "body")
    for para in body.iter(_W + "p"):
        # paragraph style id (pStyle) -> its eastAsia
        ppr = para.find(_W + "pPr")
        pstyle = None
        if ppr is not None:
            pst = ppr.find(_W + "pStyle")
            if pst is not None:
                pstyle = pst.get(_W + "val")
        para_style_ea = ea_by_style.get(pstyle, normal_ea) if pstyle else normal_ea
        for run in para.findall(_W + "r"):
            # skip runs with no text (field chars, etc.)
            texts = run.findall(_W + "t")
            if not texts or not any((t.text or "") for t in texts):
                continue
            total += 1
            # effective eastAsia: direct run rFonts, else paragraph style, else Normal.
            rpr = run.find(_W + "rPr")
            eff = None
            if rpr is not None:
                rf = rpr.find(_W + "rFonts")
                if rf is not None:
                    eff = rf.get(_W + "eastAsia")
            if eff is None:
                eff = para_style_ea
            if eff != expect_font:
                bare += 1
    check(total > 0, f"document has text runs to inspect (got {total})")
    check(bare == 0, f"no 裸 run: all {total} text runs resolve to eastAsia={expect_font} "
                     f"(offenders={bare})")

    # --- structural extras: cover shading, heading rule, PAGE-field footer ---
    doc_bytes = parts.get("word/document.xml", b"").decode("utf-8", "ignore")
    footer_bytes = "".join(
        v.decode("utf-8", "ignore") for k, v in parts.items() if "footer" in k)
    check("pBdr" in doc_bytes, "heading paragraphs carry a bottom-border rule (pBdr)")
    check(expect_title_color.upper() in doc_bytes.upper(),
          f"cover/heading deep colour {expect_title_color} present in document.xml")
    check("PAGE" in footer_bytes and "第" in footer_bytes,
          "footer carries a 「第 X 页」 PAGE field")


def main() -> int:
    biz = tokens.get_style("business")
    expect_font = biz["body_font"]              # 微软雅黑
    title_color = biz["word_title_color"]       # 1F3864 deep navy

    # ============================================================ ① Word 字体纪律
    print("== ① Word 字体纪律: styles.xml eastAsia + 字号阶梯 + 标题色 + 无裸 run ==")
    docx_path = os.path.join(_DATA, "字体纪律_business.docx")
    if os.path.exists(docx_path):
        os.remove(docx_path)
    content = (
        "# 第一章 概述\n"
        "这是正文段落，中文与 English words 混排，用于验证字体统一。\n"
        "## 1.1 小节\n"
        "- 项目符号一\n"
        "- 项目符号二\n"
        "### 细节标题\n"
        "TABLE: 季度 | 销售额 | 增长\n"
        "| 第一季度 | 1250 | 12%\n"
        "| 第二季度 | 1580 | 26%\n"
        "\n"
        "结尾正文段落。\n"
    )
    r = _FNS["write_document"](
        path=docx_path, content=content, title="年度报告",
        style="business",
        cover={"title": "如意 年度业务报告", "subtitle": "2026 财年",
               "date": "2026-07-05", "author": "把关人"},
        page_numbers=True)
    check(isinstance(r, dict) and r.get("success") is True, f"write_document ok (got {r})")
    check(r.get("style") == "business", f"reported style business (got {r.get('style')})")
    check(os.path.exists(docx_path) and zipfile.is_zipfile(docx_path), "docx is a valid OOXML zip")
    _docx_font_discipline(docx_path, expect_font, title_color)

    # ============================================================ ② Excel 落笔即样式
    print("\n== ② Excel 落笔即样式: 全表字体 + 数字格式启发式 + 数字型字符串转真数字 ==")
    xlsx_path = os.path.join(_DATA, "落笔即样式.xlsx")
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)
    rx = _FNS["write_excel"](
        path=xlsx_path,
        headers=["区域", "销售额", "增长率", "客户数"],
        data=[["华东区", "1250000", "12%", "1240"],
              ["华南区", "980000", "-5%", "890"],
              ["华北区", "2100000", "33%", "2100"]],
        style="business")
    check(isinstance(rx, dict) and rx.get("success") is True, f"write_excel ok (got {rx})")

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    # every cell (header + data) on the token font
    all_font_ok = True
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=4):
        for c in row:
            if c.value is not None and (c.font is None or c.font.name != expect_font):
                all_font_ok = False
    check(all_font_ok, f"every cell uses the token font {expect_font}")
    check(ws.cell(1, 1).font.bold is True, "header row is bold")
    # number formats
    money_fmt = ws.cell(2, 2).number_format
    pct_fmt = ws.cell(2, 3).number_format
    cnt_fmt = ws.cell(2, 4).number_format
    check("¥" in money_fmt or "#,##0.0" in money_fmt, f"销售额 col got currency format (got {money_fmt!r})")
    check("%" in pct_fmt, f"增长率 col got a percent-ish format (got {pct_fmt!r})")
    check(cnt_fmt == "#,##0", f"客户数 col got thousands format (got {cnt_fmt!r})")
    # numeric coercion
    check(isinstance(ws.cell(2, 2).value, (int, float)),
          f"numeric-looking string coerced to a real number (got {type(ws.cell(2,2).value).__name__})")

    # ============================================================ ③ 三风格 render_samples
    print("\n== ③ 三套设计系统: render_samples 三风格 × 4 件 = 12 件全部非空 ==")
    samples_dir = os.path.join(_DATA, "samples")
    # run the generator in-process by importing it
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "render_samples", os.path.join(_ROOT, "scripts", "render_samples.py"))
    rs = importlib.util.module_from_spec(spec)
    sys.argv = ["render_samples.py", samples_dir]
    spec.loader.exec_module(rs)
    rc = rs.main()
    check(rc == 0, f"render_samples exited 0 (got {rc})")
    expected = [f"sample_{st}.{ext}" for st in ("business", "minimal", "vibrant")
                for ext in ("docx", "xlsx", "pptx", "png")]
    missing = [f for f in expected
               if not (os.path.exists(os.path.join(samples_dir, f))
                       and os.path.getsize(os.path.join(samples_dir, f)) > 0)]
    check(not missing, f"all 12 sample files exist and are non-empty (missing/empty: {missing})")

    # ============================================================ ④ 向后兼容
    print("\n== ④ 向后兼容: 老式 write_document / write_excel 调用照常且自动吃新样式 ==")
    legacy_docx = os.path.join(_DATA, "legacy.docx")
    rl = _FNS["write_document"](path=legacy_docx, content="# 标题\n正文内容。\n- 要点",
                                title="旧式调用")
    check(isinstance(rl, dict) and rl.get("success") is True, f"legacy write_document ok (got {rl})")
    # legacy call defaults to business -> still gets eastAsia discipline
    with zipfile.ZipFile(legacy_docx) as z:
        styles = z.read("word/styles.xml").decode("utf-8", "ignore")
    check("eastAsia" in styles and expect_font in styles,
          "legacy docx still gets eastAsia font discipline (auto styles)")

    legacy_xlsx = os.path.join(_DATA, "legacy.xlsx")
    rlx = _FNS["write_excel"](path=legacy_xlsx, headers=["A", "B"], data=[["x", "1"], ["y", "2"]])
    check(isinstance(rlx, dict) and rlx.get("success") is True, f"legacy write_excel ok (got {rlx})")
    wb2 = load_workbook(legacy_xlsx)
    check(wb2.active.cell(1, 1).font.name == expect_font,
          "legacy xlsx cells still get the token font")

    # ---------------------------------------------------------------- summary
    print()
    if _FAILURES:
        print(f"FAILED: {len(_FAILURES)} assertion(s)")
        for f in _FAILURES:
            print("  -", f)
        print("ACC-V17 SMOKE: FAIL")
        return 1
    print("样例目录（把关人逐件 COM 导图过审美关）:")
    print("  docx/xlsx/pptx/png ×3 →", os.path.join(_DATA, "samples"))
    print("  字体纪律 docx →", docx_path)
    print("ACC-V17 SMOKE: ALL PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
