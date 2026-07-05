"""Behavioral smoke test for v1.8.0 —— 补齐「AI 盲操作」痛点的四个补充工具.

用户认可清单 (四类痛点):
  T1 excel_read     —— 结构化读表 (headers/data 二维 + 可选公式 + 数字格式 + 多 sheet 概要 + 截断)。
  T2 pdf_read_pages —— 分页读 PDF ('3'/'1-5'/'1,3,7-9')，逐页截断，尽力给大纲。
  T3 image_info / image_resize —— 图片信息 + 等比 LANCZOS 缩放 (output_path 契约 + protected 护栏)。
  T4 剪贴板文本一等公民 —— get_clipboard/set_clipboard 已具备文本读写 (本切片只补描述人话化)，本冒烟做往返验证。

本冒烟【真文件往返】断言 (不是 mock)：
  ① excel_read: 写一张带公式的 xlsx → 读回 headers/data 结构；include_formulas 读回 '=SUM(...)'；
     数字格式列被报出；多 sheet 概要；坏 range 人话错；max_rows 截断标注；中文内容原样往返。
  ② pdf_read_pages: 写一份多页 PDF → 分页读 ('1,3')；越界页人话错；max_chars_per_page 截断；中文文本读回。
  ③ image_info: 尺寸/格式/模式正确。image_resize: 等比 (只给 width 自动算 height)、scale、精确 w+h、
     JPEG (RGBA→RGB 不炸)、缺尺寸参数人话错、output_path 契约回显、protected 护栏。
  ④ 剪贴板: set_clipboard(中文) → get_clipboard 读回一致 (best-effort：无剪贴板环境则 SKIP 不 fail)。

Run with UTF-8:  python -X utf8 tests/smoke_v18.py
"""

import os
import sys
import tempfile

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(_ROOT, "src")
sys.path.insert(0, _SRC)

_DATA = os.path.join(tempfile.gettempdir(), "acc_smoke_v18_data")
os.makedirs(_DATA, exist_ok=True)
os.environ["WCW_DATA_DIR"] = _DATA

import ai_computer_control.server as server  # noqa: E402

_FNS = {t.name: t.fn for t in server.mcp._tool_manager.list_tools()}
_FAILURES: list[str] = []


def check(cond: bool, msg: str):
    print(f"  [{'ok  ' if cond else 'FAIL'}] {msg}")
    if not cond:
        _FAILURES.append(msg)


def _make_xlsx_with_formula(path: str):
    """写一张 xlsx：中文表头 + 数据 + 一列公式 + 货币数字格式 + 第二张 sheet。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "销售"
    ws.append(["区域", "单价", "数量", "金额"])
    ws["A2"], ws["B2"], ws["C2"] = "华东区", 12.5, 100
    ws["D2"] = "=B2*C2"
    ws["A3"], ws["B3"], ws["C3"] = "华南区", 8.0, 250
    ws["D3"] = "=B3*C3"
    for r in (2, 3):
        ws[f"D{r}"].number_format = "¥#,##0.00"
    ws2 = wb.create_sheet("备注")
    ws2["A1"] = "第二张表"
    wb.save(path)


def _make_pdf(path: str, pages: int):
    """用 reportlab 写一份 pages 页的中文 PDF，每页一行可辨识文本。"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        font = "STSong-Light"
    except Exception:
        font = "Helvetica"
    c = canvas.Canvas(path, pagesize=A4)
    for i in range(1, pages + 1):
        c.setFont(font, 18)
        c.drawString(72, 720, f"第 {i} 页 PAGE-MARKER-{i} 内容行")
        c.showPage()
    c.save()


def _make_png(path: str, w: int, h: int, mode="RGB"):
    from PIL import Image

    im = Image.new(mode, (w, h), (200, 120, 60) if mode == "RGB" else (200, 120, 60, 255))
    im.save(path)


def main() -> int:
    # ============================================================ ① excel_read
    print("== ① excel_read: 结构化读表 / 公式 / 数字格式 / 多 sheet / 坏 range / 截断 / 中文往返 ==")
    xp = os.path.join(_DATA, "结构化读表.xlsx")
    _make_xlsx_with_formula(xp)

    r = _FNS["excel_read"](path=xp)
    check(r.get("ok") is True, f"excel_read 基本读 ok (got {r.get('error')})")
    check(r.get("sheet") == "销售", f"缺省读活动表「销售」(got {r.get('sheet')!r})")
    check(r.get("headers") == ["区域", "单价", "数量", "金额"],
          f"headers 结构化读回 (got {r.get('headers')!r})")
    check(len(r.get("data", [])) == 2 and r["data"][0][0] == "华东区",
          f"data 二维读回，首行中文原样 (got {r.get('data')!r})")
    # 真数字列原样读回 (单价 12.5 是真数字，非公式)。
    check(r["data"][0][1] == 12.5, f"真数字单元格原样读回 12.5 (got {r['data'][0][1]!r})")
    # 公式列: data_only=True 取【缓存的算好值】。本 xlsx 由 openpyxl 写、从未被 Excel 打开过 → 公式无缓存值
    #   → 读回 None (openpyxl 的既知行为，非 excel_read 的 bug；真 Excel 存的文件会带缓存值)。include_formulas
    #   分支才是拿公式串的正道 (下面验证)。这里断言「值分支对无缓存公式返回 None」这一诚实事实。
    check(r["data"][0][3] is None,
          f"无缓存公式格在值模式读回 None (openpyxl 既知行为) (got {r['data'][0][3]!r})")
    # 多 sheet 概要
    names = {s["name"] for s in r.get("sheets", [])}
    check(names == {"销售", "备注"}, f"sheets 概要含两张表 (got {names})")
    # 数字格式列 (D 列 = 金额, ¥#,##0.00)
    nf = r.get("number_formats") or {}
    check(nf.get("D") == "¥#,##0.00", f"D 列数字格式被报出 (got {nf})")

    # include_formulas —— 公式读回 (本切片核心断言之一)
    rf = _FNS["excel_read"](path=xp, include_formulas=True)
    fm = rf.get("formulas") or {}
    check(fm.get("D2") == "=B2*C2" and fm.get("D3") == "=B3*C3",
          f"include_formulas 读回公式串 (got {fm})")

    # 值往返: 用 ACC 自己的 write_excel 写 (数字型字符串被转成真数字 → 有真实值可读回)，验证 data 值链路。
    xp_vals = os.path.join(_DATA, "值往返.xlsx")
    _FNS["write_excel"](path=xp_vals, headers=["季度", "销售额"],
                        data=[["Q1", "1250"], ["Q2", "1580"]], style="business")
    rv = _FNS["excel_read"](path=xp_vals)
    check(rv.get("ok") is True and rv["data"][0] == ["Q1", 1250] and rv["data"][1] == ["Q2", 1580],
          f"write_excel→excel_read 值往返 (数字真值读回) (got {rv.get('data')!r})")

    # 指定 sheet + range
    rr = _FNS["excel_read"](path=xp, sheet="销售", range="A1:B2")
    check(rr.get("ok") is True and rr.get("range") == "A1:B2", f"指定 range 生效 (got {rr.get('range')!r})")
    check(rr.get("headers") == ["区域", "单价"], f"range 限定列 (got {rr.get('headers')!r})")

    # 坏 range → 人话错
    bad = _FNS["excel_read"](path=xp, range="不是区域!!")
    check(bad.get("error") and "区域" in bad["error"], f"坏 range 人话错 (got {bad.get('error')!r})")

    # 缺 sheet → 人话错列出可选
    bad2 = _FNS["excel_read"](path=xp, sheet="不存在")
    check(bad2.get("error") and "销售" in bad2["error"], f"坏 sheet 名人话错列出可选 (got {bad2.get('error')!r})")

    # max_rows 截断标注
    rt = _FNS["excel_read"](path=xp, max_rows=1)
    tr = rt.get("truncated")
    check(isinstance(tr, dict) and tr.get("rows_total") == 3 and tr.get("rows_returned") == 1,
          f"max_rows=1 截断标注真实总行数 (got {tr})")

    # ============================================================ ② pdf_read_pages
    print("\n== ② pdf_read_pages: 分页 / 越界人话错 / 每页截断 / 中文读回 / 大纲字段 ==")
    pp = os.path.join(_DATA, "分页读.pdf")
    _make_pdf(pp, pages=4)

    pr = _FNS["pdf_read_pages"](path=pp, pages="1,3")
    check(pr.get("ok") is True, f"pdf_read_pages ok (got {pr.get('error')})")
    check(pr.get("total_pages") == 4, f"total_pages=4 (got {pr.get('total_pages')})")
    got_pages = [p["page"] for p in pr.get("pages", [])]
    check(got_pages == [1, 3], f"只读点名的页 [1,3] (got {got_pages})")
    check("PAGE-MARKER-1" in pr["pages"][0]["text"] and "PAGE-MARKER-3" in pr["pages"][1]["text"],
          "读回的正是第 1、3 页的文本 (页码映射正确)")
    check("内容行" in pr["pages"][0]["text"], "中文文本读回")
    check("outline" in pr and isinstance(pr["outline"], list), "outline 字段存在 (无 pypdf 时为 [])")

    # 区间 + 混合
    pr2 = _FNS["pdf_read_pages"](path=pp, pages="2-4")
    check([p["page"] for p in pr2["pages"]] == [2, 3, 4], "区间 '2-4' 解析正确")

    # 越界 → 人话错
    pbad = _FNS["pdf_read_pages"](path=pp, pages="3-9")
    check(pbad.get("error") and "越界" in pbad["error"] and "共 4 页" in pbad["error"],
          f"越界页人话错 (got {pbad.get('error')!r})")

    # 每页截断
    pt = _FNS["pdf_read_pages"](path=pp, pages="1", max_chars_per_page=5)
    check(pt["pages"][0]["truncated"] is True and pt["pages"][0]["chars"] == 5,
          f"max_chars_per_page 截断 (got chars={pt['pages'][0]['chars']} trunc={pt['pages'][0]['truncated']})")

    # ============================================================ ③ image_info / image_resize
    print("\n== ③ image_info / image_resize: 信息 / 等比 / scale / 精确 / JPEG / 缺参 / 契约 / 护栏 ==")
    ip = os.path.join(_DATA, "原图.png")
    _make_png(ip, 400, 200, mode="RGB")

    ii = _FNS["image_info"](path=ip)
    check(ii.get("ok") is True and ii.get("width") == 400 and ii.get("height") == 200,
          f"image_info 尺寸正确 (got {ii.get('width')}x{ii.get('height')})")
    check(ii.get("format") == "PNG" and ii.get("mode") == "RGB", f"image_info 格式/模式 (got {ii.get('format')}/{ii.get('mode')})")
    check(isinstance(ii.get("file_size"), int) and ii["file_size"] > 0, "image_info 文件大小 > 0")

    # 等比：只给 width → height 自动 = 200 * (100/400) = 50
    op1 = os.path.join(_DATA, "缩_等比.png")
    q1 = _FNS["image_resize"](path=ip, output_path=op1, width=100)
    check(q1.get("ok") is True and q1.get("new_size") == [100, 50],
          f"只给 width 等比算高 → [100,50] (got {q1.get('new_size')})")
    check(q1.get("output_path") == os.path.abspath(op1), "image_resize output_path 契约回显 (== 落盘绝对路径)")
    # 真验证落盘尺寸
    v1 = _FNS["image_info"](path=op1)
    check(v1.get("width") == 100 and v1.get("height") == 50, f"落盘图真实尺寸 100x50 (got {v1.get('width')}x{v1.get('height')})")

    # scale = 0.25 → 100x50
    op2 = os.path.join(_DATA, "缩_scale.png")
    q2 = _FNS["image_resize"](path=ip, output_path=op2, scale=0.25)
    check(q2.get("new_size") == [100, 50], f"scale=0.25 → [100,50] (got {q2.get('new_size')})")

    # 精确 w+h（可变形）
    op3 = os.path.join(_DATA, "缩_精确.png")
    q3 = _FNS["image_resize"](path=ip, output_path=op3, width=80, height=80)
    check(q3.get("new_size") == [80, 80], f"精确 w+h → [80,80] (got {q3.get('new_size')})")

    # RGBA 源 → JPEG 输出（不炸，自动转 RGB）
    ipa = os.path.join(_DATA, "原图_rgba.png")
    _make_png(ipa, 300, 300, mode="RGBA")
    op4 = os.path.join(_DATA, "缩.jpg")
    q4 = _FNS["image_resize"](path=ipa, output_path=op4, width=150, quality=80)
    check(q4.get("ok") is True and q4.get("format") == "JPEG", f"RGBA→JPEG 输出不炸 (got {q4})")
    v4 = _FNS["image_info"](path=op4)
    check(v4.get("format") == "JPEG" and v4.get("mode") == "RGB", f"JPEG 落盘 mode=RGB (got {v4.get('mode')})")

    # 缺尺寸参数 → 人话错
    qbad = _FNS["image_resize"](path=ip, output_path=os.path.join(_DATA, "x.png"))
    check(qbad.get("error") and "width" in qbad["error"], f"缺尺寸参数人话错 (got {qbad.get('error')!r})")

    # protected 护栏：往 SystemRoot 写 → refused
    sysroot = os.environ.get("SystemRoot", r"C:\Windows")
    qprot = _FNS["image_resize"](path=ip, output_path=os.path.join(sysroot, "wcw_should_refuse.png"), width=50)
    check(qprot.get("error") and "refused" in qprot["error"], f"protected 护栏拦截 (got {qprot.get('error')!r})")

    # 源不存在 → 人话错
    qmiss = _FNS["image_resize"](path=os.path.join(_DATA, "没有这个.png"),
                                 output_path=os.path.join(_DATA, "y.png"), width=50)
    check(qmiss.get("error") and "不存在" in qmiss["error"], "源不存在人话错")

    # ============================================================ ④ 剪贴板文本往返 (best-effort)
    print("\n== ④ 剪贴板文本一等公民: set_clipboard(中文) → get_clipboard 往返 (无环境则 SKIP) ==")
    marker = "如意剪贴板往返_UNIQUE_中文测试_✓"
    sc = _FNS["set_clipboard"](text=marker)
    if sc.get("ok"):
        gc = _FNS["get_clipboard"]()
        if gc.get("ok") and isinstance(gc.get("text"), str):
            check(gc["text"] == marker, f"剪贴板中文往返一致 (got {gc.get('text')!r})")
        else:
            print("  [skip] get_clipboard 读不到 (无剪贴板后端) —— 不计失败")
    else:
        print(f"  [skip] set_clipboard 无剪贴板环境 ({sc.get('error')}) —— 不计失败")
    check(sc.get("ok") is True or "error" in sc, "set_clipboard 返回规范信封 (ok 或 error)")

    # ---------------------------------------------------------------- summary
    print()
    if _FAILURES:
        print(f"FAILED: {len(_FAILURES)} assertion(s)")
        for f in _FAILURES:
            print("  -", f)
        print("ACC-V18 SMOKE: FAIL")
        return 1
    print("产物 (真文件往返):")
    for f in ("结构化读表.xlsx", "分页读.pdf", "原图.png", "缩_等比.png", "缩.jpg"):
        print("  ", os.path.join(_DATA, f))
    print("ACC-V18 SMOKE: ALL PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
