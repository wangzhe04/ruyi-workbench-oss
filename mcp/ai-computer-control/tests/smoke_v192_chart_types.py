"""Behavioral smoke test for 109c — ACC 出图补型 (Python 侧).

覆盖:
  ① chart_image 新增 3 型 (hbar/stacked_bar/area)，连同既有 4 型 (bar/line/pie/scatter) 共 7 型，
     全部能用一份 2 系列 CJK 数据集渲染出非空 PNG (pie 依既有校验只能单系列，单独给一份单系列数据)。
  ② chart_image 未知 chart_type 返回失败 (无 success:true)，错误信息里列出全部 7 个合法型名。
  ③ excel_chart 新增 scatter (openpyxl ScatterChart)：首列数值 X + 其余每列一条 Y 系列；落盘后重新
     加载断言恰有一个 scatterChart 且系列数与 Y 列数一致。
  ④ excel_chart 既有 bar 型不受影响 (向后兼容对照)。

matplotlib / openpyxl 均为可选离线依赖 —— 任一缺失则打印 SKIP 并 exit 0，不把环境问题误报成回归
(镜像 office_chart.py / office_excel.py 自身的降级处理)。

Run with UTF-8:  python -X utf8 tests/smoke_v192_chart_types.py
"""

import os
import sys
import tempfile

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(_ROOT, "src")
sys.path.insert(0, _SRC)

_DATA = os.path.join(tempfile.gettempdir(), "acc_smoke_v192_data")
os.makedirs(_DATA, exist_ok=True)
os.environ.setdefault("WCW_DATA_DIR", _DATA)

try:
    import matplotlib  # noqa: F401
    import openpyxl  # noqa: F401
    _DEPS_OK = True
except Exception:
    _DEPS_OK = False

_FAILURES: list[str] = []

# The 7 chart_type values chart_image must support after 109c.
_CHART_TYPES = ("bar", "line", "pie", "scatter", "hbar", "stacked_bar", "area")


def check(cond: bool, msg: str):
    print(f"  [{'ok  ' if cond else 'FAIL'}] {msg}")
    if not cond:
        _FAILURES.append(msg)


def main() -> int:
    if not _DEPS_OK:
        print("SKIP: matplotlib/openpyxl 不可用，跳过 109c 出图补型冒烟（可选离线依赖）")
        return 0

    import ai_computer_control.server as server  # noqa: E402

    _FNS = {t.name: t.fn for t in server.mcp._tool_manager.list_tools()}

    print("# 109c smoke: chart_image 补型 (hbar/stacked_bar/area) + excel_chart scatter")

    # ============================================================ ① chart_image 七型全渲染
    print("\n== ① chart_image: 7 型全部可渲染出非空 PNG (CJK 数据集) ==")
    multi = {
        "labels": ["一月", "二月", "三月"],
        "series": [
            {"name": "销售额", "values": [120, 150, 90]},
            {"name": "成本", "values": [80, 100, 70]},
        ],
    }
    single = {
        "labels": ["一月", "二月", "三月"],
        "series": [{"name": "销售额", "values": [120, 150, 90]}],
    }
    for ctype in _CHART_TYPES:
        # 饼图受既有校验约束只能单系列，其余型用 2 系列数据集。
        d = single if ctype == "pie" else multi
        p = os.path.join(_DATA, f"chart_{ctype}.png")
        r = _FNS["chart_image"](path=p, chart_type=ctype, data=d, title=f"测试图-{ctype}")
        check(r.get("success") is True, f"{ctype}: success (got error={r.get('error')})")
        b = r.get("bytes")
        check(isinstance(b, int) and b > 1000, f"{ctype}: bytes > 1000 (got {b!r})")
        check(r.get("chart_type") == ctype, f"{ctype}: chart_type 回显一致 (got {r.get('chart_type')!r})")
        check(os.path.exists(p) and os.path.getsize(p) > 1000, f"{ctype}: PNG 文件落盘且非空")

    # ============================================================ ② 未知 chart_type
    print("\n== ② chart_image: 未知 chart_type 失败且错误信息列出 7 个合法型名 ==")
    r_bad = _FNS["chart_image"](path=os.path.join(_DATA, "chart_bad.png"), chart_type="donut",
                                data=multi, title="非法型测试")
    check(r_bad.get("success") is not True, f"未知型不返回 success (got {r_bad})")
    err = str(r_bad.get("error", ""))
    check(bool(err), "未知型返回 error 说明")
    for t in _CHART_TYPES:
        check(t in err, f"错误信息含合法型名 '{t}' (got {err!r})")

    # ============================================================ ③ excel_chart scatter
    print("\n== ③ excel_chart: scatter (ScatterChart, 首列 X + 每列一条 Y 系列) ==")
    from openpyxl import Workbook, load_workbook

    xp = os.path.join(_DATA, "散点测试.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["X", "Y1", "Y2"])
    for i in range(6):
        ws.append([i + 1, (i + 1) * 2, (i + 1) * 3 - 1])
    wb.save(xp)

    r_sc = _FNS["excel_chart"](path=xp, sheet="Sheet1", chart_type="scatter",
                               data_range="A1:C7", title="散点测试")
    check(r_sc.get("success") is True, f"scatter: success (got error={r_sc.get('error')})")
    check(r_sc.get("chart_type") == "scatter", f"scatter: chart_type 回显 (got {r_sc.get('chart_type')!r})")

    wb2 = load_workbook(xp)
    scatter_charts = [c for c in wb2.active._charts if c.tagname == "scatterChart"]
    check(len(scatter_charts) == 1, f"scatter: 恰有一个 scatterChart (got {len(scatter_charts)})")
    if scatter_charts:
        n_series = len(scatter_charts[0].series)
        check(n_series == 2, f"scatter: 2 条 Y 系列 (data_range 3 列 = 1 列 X + 2 列 Y) (got {n_series})")

    # ============================================================ ④ excel_chart bar 向后兼容
    print("\n== ④ excel_chart: bar 型不受影响 (向后兼容对照) ==")
    r_bar = _FNS["excel_chart"](path=xp, sheet="Sheet1", chart_type="bar",
                                data_range="A1:C7", title="柱状对照", target_cell="M2")
    check(r_bar.get("success") is True, f"bar: success (got error={r_bar.get('error')})")
    wb3 = load_workbook(xp)
    bar_charts = [c for c in wb3.active._charts if c.tagname == "barChart"]
    check(len(bar_charts) == 1, f"bar: 恰有一个 barChart (got {len(bar_charts)})")

    # ============================================================ summary
    print("\n" + ("=" * 60))
    if _FAILURES:
        print(f"FAILED: {len(_FAILURES)} assertion(s)")
        for f in _FAILURES:
            print("  -", f)
        return 1
    print("OK: all 109c chart-type assertions passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
