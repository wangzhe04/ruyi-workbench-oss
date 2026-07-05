"""结构化读表 / 分页读 PDF —— 解决「AI 盲操作」类痛点 (v1.8.0).

read_document (document.py) 的 xlsx/pdf 分支保留不动 —— 那是「把整篇拍平成一段文本」的粗读，适合小文件
快速一瞥。本模块是【增强版结构化读】:

  * excel_read      —— 结构化读表: 返回 headers/data 二维、可选公式、数字格式、多 sheet 概要，
                        read_only 防大文件卡死，max_rows 截断防上下文爆掉。
  * pdf_read_pages  —— 分页读 PDF: 只读你点名的页 (支持 '3' / '1-5' / '1,3,7-9')，逐页截断，
                        试给大纲 (outline)。解决「50 页 PDF 整读爆上下文」。

依赖纪律: openpyxl / pdfplumber 都在核心依赖树里 (requirements_offline.txt)。若某台机器缺失，工具返回
人话安装提示而非炸服务器启动 (import 守护)。pypdf 不在依赖树 —— pdf 大纲能拿到就给、拿不到就如实说明。
"""

import os

from ai_computer_control.server import mcp


# =============================================================================================
# T1 excel_read —— 结构化读表
# =============================================================================================
def _col_letter(idx: int) -> str:
    """1-based 列号 → Excel 列字母 (1→A, 27→AA)。"""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def _parse_a1_range(rng: str):
    """把 'B2:D10' 拆成 (min_row, min_col, max_row, max_col)，1-based。坏 range 抛 ValueError (人话)。"""
    from openpyxl.utils import range_boundaries  # (min_col, min_row, max_col, max_row)

    try:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng).strip())
    except Exception:
        raise ValueError(
            f"看不懂的单元格区域 range={rng!r} —— 请用 A1 记法，例如 'A1:D20' 或单格 'B2'。"
        )
    if None in (min_col, min_row, max_col, max_row):
        raise ValueError(
            f"不完整的单元格区域 range={rng!r} —— 需要形如 'A1:D20' 的闭区间 (不支持整列 'A:A')。"
        )
    return min_row, min_col, max_row, max_col


@mcp.tool()
def excel_read(
    path: str,
    sheet: str | None = None,
    range: str | None = None,
    include_formulas: bool = False,
    max_rows: int = 200,
) -> dict:
    """结构化读取 Excel (.xlsx) —— 让 AI 看清表结构，不再盲操作。

    这是 read_document 的【增强版结构化读】(那个粗读把整篇拍平成一段文本；本工具返回可直接遍历的二维
    data + 表头 + 可选公式/数字格式)。多 sheet 时缺省返回「各 sheet 概要 + active sheet 的数据」，你可再
    指定 sheet= 精读某张。

    Args:
        path: .xlsx 文件路径。
        sheet: 要读的工作表名；缺省读活动表 (active sheet)，并在 sheets 里给出全部表的概要。
        range: A1 记法的区域 (如 'A1:D50' / 单格 'B2')；缺省从 A1 起读到数据末尾 (受 max_rows 限)。
        include_formulas: True 时【额外】回读公式 (如 {'D2':'=SUM(B2:C2)'}) —— 需再开一次工作簿
            (data_only=False)，稍慢；缺省 False 只给算好的值。
        max_rows: 数据行数上限 (防大表爆上下文)，默认 200。超出则截断并在 truncated 标注真实总行数。

    Returns:
        dict with:
          ok, path, sheet (实际读的表名),
          sheets: [{name, rows, cols}, ...]  —— 全部工作表的维度概要,
          headers: [...]           —— 第一读入行当表头 (纯展示用；data 不含它),
          data: [[...], ...]       —— 二维单元格值 (字符串/数字/None)，已按 max_rows 截断,
          truncated: {rows_returned, rows_total, note} | None  —— 截断时给出,
          formulas: {'A1':'=...'} | None   —— include_formulas 时给出 (仅含真的有公式的格),
          number_formats: {'B':'¥#,##0.00', ...} | None  —— 有非通用数字格式的列 (列字母→格式串),
          range: 实际读取的 A1 区域。
        缺 openpyxl / 文件不存在 / 坏 range / 坏 sheet 名 → {'error': 人话说明}。
    """
    if not str(path).lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return {"error": f"excel_read 只读 .xlsx 系列，收到 {path!r}。CSV/老 .xls 不支持。"}
    if not os.path.exists(path):
        return {"error": f"文件不存在: {path}"}

    try:
        from openpyxl import load_workbook
    except Exception:
        return {"error": "结构化读表需要 openpyxl。离线包已含，可运行 installer 重装；或 pip install openpyxl"}

    try:
        mr = max(1, int(max_rows))
    except Exception:
        mr = 200

    try:
        # read_only=True: 流式读，防大文件把内存/时间吃爆 (用户明确要求)。data_only=True: 取算好的值。
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"打不开工作簿 (可能损坏或非 xlsx): {e}"}

    try:
        # 全表概要 (维度) —— 多 sheet 时让 AI 一眼看清有哪些表、各多大。
        sheets_summary = []
        for nm in wb.sheetnames:
            w = wb[nm]
            sheets_summary.append({
                "name": nm,
                "rows": int(w.max_row or 0),
                "cols": int(w.max_column or 0),
            })

        # 选表: 显式 sheet= 优先，否则活动表。
        if sheet is not None:
            if sheet not in wb.sheetnames:
                wb.close()
                return {"error": f"没有名为 {sheet!r} 的工作表。可选: {wb.sheetnames}"}
            ws = wb[sheet]
            active_name = sheet
        else:
            ws = wb.active
            active_name = ws.title

        # 区域: 显式 range= 优先，否则整表 (A1 起)。
        if range is not None:
            try:
                min_row, min_col, max_row, max_col = _parse_a1_range(range)
            except ValueError as ve:
                wb.close()
                return {"error": str(ve)}
        else:
            min_row, min_col = 1, 1
            max_row = int(ws.max_row or 1)
            max_col = int(ws.max_column or 1)

        rows_total = max_row - min_row + 1
        row_cap = min(rows_total, mr)
        hard_max_row = min_row + row_cap - 1

        # 读值 (逐行，read_only 迭代器)。
        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=hard_max_row,
                                min_col=min_col, max_col=max_col, values_only=True):
            data.append([cell for cell in row])

        headers = [("" if v is None else v) for v in data[0]] if data else []
        body = data[1:] if len(data) > 1 else []

        truncated = None
        if rows_total > row_cap:
            truncated = {
                "rows_returned": row_cap,
                "rows_total": rows_total,
                "note": f"数据超 max_rows={mr}，只返回前 {row_cap} 行 (共 {rows_total} 行)。"
                        f"缩小 range= 或调大 max_rows= 读更多。",
            }

        # 数字格式: read_only 模式的 cell 不带 number_format，需要普通模式补一遍 (只扫第一数据行的每列)。
        number_formats = _column_number_formats(path, active_name, min_row, min_col, max_col, hard_max_row)

        wb.close()

        out = {
            "ok": True,
            "path": os.path.abspath(path),
            "sheet": active_name,
            "sheets": sheets_summary,
            "headers": headers,
            "data": body,
            "range": f"{_col_letter(min_col)}{min_row}:{_col_letter(max_col)}{hard_max_row}",
            "truncated": truncated,
            "number_formats": number_formats or None,
            "formulas": None,
        }

        # include_formulas: 单独开一次 data_only=False (values 变成公式串)，只挑真有公式的格。
        if include_formulas:
            out["formulas"] = _read_formulas(path, active_name, min_row, min_col, max_row, max_col, mr)

        return out
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"error": f"读表失败: {type(e).__name__}: {e}"}


def _column_number_formats(path, sheet_name, min_row, min_col, max_col, max_row) -> dict:
    """扫每列首个数据格的 number_format，非 'General' 则记 (列字母→格式串)。read_only 拿不到格式，故普通开。"""
    out = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb[sheet_name]
        # 从表头下一行 (若有) 起找首个非空格取格式；单行区域就用该行。
        probe_start = min_row + 1 if max_row > min_row else min_row
        for col in range(min_col, max_col + 1):
            fmt = None
            for r in range(probe_start, max_row + 1):
                c = ws.cell(row=r, column=col)
                nf = c.number_format
                if nf and nf != "General":
                    fmt = nf
                    break
            if fmt:
                out[_col_letter(col)] = fmt
        wb.close()
    except Exception:
        return {}
    return out


def _read_formulas(path, sheet_name, min_row, min_col, max_row, max_col, mr) -> dict:
    """回读区域内的公式 (data_only=False → cell.value 是 '=...' 串)。仅收真以 '=' 开头的格。"""
    out = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb[sheet_name]
        hard_max_row = min_row + min(max_row - min_row + 1, mr) - 1
        r = min_row - 1
        for row in ws.iter_rows(min_row=min_row, max_row=hard_max_row,
                                min_col=min_col, max_col=max_col, values_only=False):
            r += 1
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    out[f"{_col_letter(cell.column)}{cell.row}"] = v
        wb.close()
    except Exception:
        return {}
    return out


# =============================================================================================
# T2 pdf_read_pages —— 分页读 PDF
# =============================================================================================
def _parse_pages(spec: str, total: int) -> list[int]:
    """'3' / '1-5' / '1,3,7-9' → 有序去重的 1-based 页号列表。越界/坏格式抛 ValueError (人话)。"""
    spec = str(spec).strip()
    if not spec:
        raise ValueError("pages 不能为空 —— 例如 '3' / '1-5' / '1,3,7-9'。")
    pages: list[int] = []
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            lo_s, _, hi_s = part.partition("-")
            try:
                lo, hi = int(lo_s), int(hi_s)
            except ValueError:
                raise ValueError(f"看不懂的页码段 {part!r} —— 例如 '1-5'。")
            if lo > hi:
                lo, hi = hi, lo
            for p in range(lo, hi + 1):
                pages.append(p)
        else:
            try:
                pages.append(int(part))
            except ValueError:
                raise ValueError(f"看不懂的页码 {part!r} —— 例如 '3'。")
    if not pages:
        raise ValueError(f"pages={spec!r} 没解析出任何页码。")
    bad = sorted({p for p in pages if p < 1 or p > total})
    if bad:
        raise ValueError(
            f"页码越界: {bad} —— 本 PDF 共 {total} 页 (有效页码 1..{total})。"
        )
    # 去重保序。
    seen, ordered = set(), []
    for p in pages:
        if p not in seen:
            seen.add(p)
            ordered.append(p)
    return ordered


def _pdf_outline(path: str) -> list[dict]:
    """尽力取 PDF 大纲/书签 [{title, page}]。pdfplumber 拿不到大纲，试 pypdf；pypdf 不在依赖树则返回 []。"""
    try:
        import pypdf  # 不在依赖树 —— 装了就用，没装就静默回 []。
    except Exception:
        return []
    try:
        reader = pypdf.PdfReader(path)
        out = []

        def _walk(items):
            for it in items:
                if isinstance(it, list):
                    _walk(it)
                    continue
                try:
                    title = str(getattr(it, "title", "") or "")
                    pageno = reader.get_destination_page_number(it) + 1  # 0-based → 1-based
                    out.append({"title": title, "page": int(pageno)})
                except Exception:
                    continue

        _walk(reader.outline or [])
        return out
    except Exception:
        return []


@mcp.tool()
def pdf_read_pages(path: str, pages: str = "1-5", max_chars_per_page: int = 4000) -> dict:
    """分页读取 PDF —— 只读你点名的页，解决「50 页 PDF 整读爆上下文」。

    read_document 会把整本 PDF 拍平成一大段文本 (小文件够用)；本工具让你精确点页 (支持 '3' / '1-5' /
    '1,3,7-9')，逐页截断，并尽力给出大纲让你先看目录再决定读哪页。

    Args:
        path: .pdf 文件路径。
        pages: 页码规格，1-based。支持 '3' (单页) / '1-5' (区间) / '1,3,7-9' (混合)。默认 '1-5'。
        max_chars_per_page: 每页文本上限 (防单页超长爆上下文)，默认 4000。超出截断并 truncated=True。

    Returns:
        dict with:
          ok, path, total_pages,
          outline: [{title, page}, ...]  —— PDF 书签大纲；pdfplumber 拿不到，若装了 pypdf 则用它，
                   否则为 [] 并在 outline_note 说明「未装 pypdf，无法读大纲」,
          outline_note: str | None,
          pages: [{page, text, chars, truncated}, ...]  —— 按请求顺序，每页的文本。
        越界页码 / 坏 pages 格式 / 缺 pdfplumber / 文件不存在 → {'error': 人话说明}。
    """
    if not str(path).lower().endswith(".pdf"):
        return {"error": f"pdf_read_pages 只读 .pdf，收到 {path!r}。"}
    if not os.path.exists(path):
        return {"error": f"文件不存在: {path}"}

    try:
        import pdfplumber
    except Exception:
        return {"error": "分页读 PDF 需要 pdfplumber。离线包已含，可运行 installer 重装；或 pip install pdfplumber"}

    try:
        cap = max(1, int(max_chars_per_page))
    except Exception:
        cap = 4000

    try:
        with pdfplumber.open(path) as pdf:
            total = len(pdf.pages)
            try:
                want = _parse_pages(pages, total)
            except ValueError as ve:
                return {"error": str(ve)}

            out_pages = []
            for pno in want:
                page = pdf.pages[pno - 1]  # 1-based → 0-based
                text = page.extract_text() or ""
                truncated = False
                if len(text) > cap:
                    text = text[:cap]
                    truncated = True
                out_pages.append({
                    "page": pno,
                    "text": text,
                    "chars": len(text),
                    "truncated": truncated,
                })

        outline = _pdf_outline(path)
        outline_note = None
        if not outline:
            outline_note = "未读到大纲 (PDF 无书签，或未安装 pypdf —— pypdf 非默认依赖，装后可读大纲)。"

        return {
            "ok": True,
            "path": os.path.abspath(path),
            "total_pages": total,
            "outline": outline,
            "outline_note": outline_note,
            "pages": out_pages,
        }
    except Exception as e:
        return {"error": f"读 PDF 失败: {type(e).__name__}: {e}"}
