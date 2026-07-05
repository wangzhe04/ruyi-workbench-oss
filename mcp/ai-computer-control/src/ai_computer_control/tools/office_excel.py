"""Excel beautify + chart tools (v1.6 模板驱动).

excel_beautify — turn a plain data .xlsx (just written by write_excel) into a professional-looking
                 sheet: bold white header on the primary fill, frozen header row, zebra striping,
                 thin borders over the used range, content-fit column widths (cap 50), right-aligned
                 numeric columns, auto-filter. Idempotent — re-running does not stack/duplicate styles.
excel_chart    — insert a native openpyxl chart (bar|line|pie) at a target cell, coloured from the
                 design tokens.

Both use openpyxl (already a core dependency, so no import guard needed for the happy path — but we
still degrade to a 中文人话 error if openpyxl is somehow unimportable, to honour the "never let the
MCP fail to start" discipline).
"""

import os

from ai_computer_control.server import mcp
from ai_computer_control.tools.safety import protected_path_reason
from ai_computer_control.tools import office_style as style_tokens


def _protected_write_guard(path: str, allow_protected: bool):
    """Mirror document.py's write guard: refuse writing into a protected system tree unless overridden."""
    reason = protected_path_reason(path)
    if reason and not allow_protected:
        return {"error": f"拒绝写入：目标 {reason}。如确需写入，请传 allow_protected=true。"}
    return None


def _require_openpyxl():
    """Import openpyxl pieces lazily; return (modules_dict, None) or (None, error_dict)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        return {
            "load_workbook": load_workbook,
            "Font": Font,
            "PatternFill": PatternFill,
            "Alignment": Alignment,
            "Border": Border,
            "Side": Side,
            "get_column_letter": get_column_letter,
        }, None
    except Exception as e:  # noqa: BLE001
        return None, {"error": f"Excel 美化需要 openpyxl（应随离线包安装）。导入失败：{e}"}


def _is_number(value) -> bool:
    """True if the cell value is a real number (int/float, not bool)."""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


@mcp.tool(audit=True)
def excel_beautify(
    path: str,
    sheet: str | None = None,
    style: str = "business",
    header_row: int = 1,
    allow_protected: bool = False,
) -> dict:
    """Beautify an existing .xlsx into a professional-looking sheet (模板驱动, idempotent).

    Applies, using the chosen design-token style: a bold header row (business/vibrant: primary fill
    + white text; minimal「墨白极简」v1.7.1: 白底墨黑粗体 + 底部 2pt 青色下边框 — 不用满底色), a frozen
    header row, zebra striping on the data rows (minimal 的斑马纹为极浅灰), thin borders over the whole
    used range, content-fit column widths (capped at 50 chars), right-aligned numeric columns, and an
    auto-filter over the used range. Re-running is safe — styling is recomputed from the data, not
    layered on top, so it never accumulates.

    Args:
        path: Path to an existing .xlsx (e.g. just written by write_excel).
        sheet: Worksheet name; None = the active/first sheet.
        style: Design style — 'business' (default) | 'minimal' | 'vibrant'. Unknown -> 'business'.
        header_row: 1-based row index of the header row (default 1). Rows below it are data.
        allow_protected: Override the protected-system-root guard on the destination (default off).

    Returns:
        dict with 'success', 'path', 'output_path', 'sheet', 'style', 'rows', 'cols'. On failure a
        {'error': <中文人话>} dict (missing file / bad sheet / no data / import failure).
    """
    if not str(path).lower().endswith(".xlsx"):
        return {"error": "path 必须以 .xlsx 结尾"}
    if not os.path.exists(path):
        return {"error": f"文件不存在：{path}"}
    guard = _protected_write_guard(path, allow_protected)
    if guard:
        return guard
    mods, err = _require_openpyxl()
    if err:
        return err

    try:
        header_row = int(header_row)
        if header_row < 1:
            return {"error": "header_row 必须 >= 1"}
    except (TypeError, ValueError):
        return {"error": "header_row 必须是整数"}

    tokens = style_tokens.get_style(style)
    resolved_style = style if style in style_tokens.STYLES else style_tokens.DEFAULT_STYLE

    Font = mods["Font"]
    PatternFill = mods["PatternFill"]
    Alignment = mods["Alignment"]
    Border = mods["Border"]
    Side = mods["Side"]
    get_column_letter = mods["get_column_letter"]

    try:
        wb = mods["load_workbook"](path)
        if sheet is not None:
            if sheet not in wb.sheetnames:
                wb.close()
                return {"error": f"工作表不存在：{sheet}。可用：{wb.sheetnames}"}
            ws = wb[sheet]
        else:
            ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < header_row or max_col < 1:
            wb.close()
            return {"error": "工作表没有可美化的数据（为空或表头行超出范围）"}

        # --- shared style objects (build once) ---
        # v1.7.1 表头版式选择器: 'fill' (business/vibrant) = 主色满底白粗字 (v1.6 原样);
        # 'underline' (minimal「墨白极简」) = 白底墨黑粗字 + 底部 2pt 青色下边框 (不用满底色).
        header_mode = tokens.get("excel_header", "fill")
        zebra_fill = PatternFill(fill_type="solid", fgColor=style_tokens.argb(tokens["zebra_fill"]))
        no_fill = PatternFill(fill_type=None)  # explicit clear -> idempotent (strips a prior zebra tint)
        body_font = Font(name=tokens["body_font"], bold=False,
                         color=style_tokens.argb(tokens["text_color"]))
        thin = Side(style="thin", color=style_tokens.argb(tokens["border_color"]))
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if header_mode == "underline":
            header_fill = no_fill   # 白底 (explicit clear keeps re-runs idempotent)
            header_font = Font(name=tokens["body_font"], bold=True,
                               color=style_tokens.argb(tokens["text_color"]))
            # 'medium' ≈ 2pt in Excel border weights; 青色 accent line under the header row.
            accent_bottom = Side(style="medium", color=style_tokens.argb(tokens["accent"]))
            header_border = Border(left=thin, right=thin, top=thin, bottom=accent_bottom)
        else:
            header_fill = PatternFill(fill_type="solid",
                                      fgColor=style_tokens.argb(tokens["header_fill"]))
            header_font = Font(name=tokens["body_font"], bold=True,
                               color=style_tokens.argb(tokens["header_font_color"]))
            header_border = border
        left_mid = Alignment(horizontal="left", vertical="center")
        right_mid = Alignment(horizontal="right", vertical="center")
        center_mid = Alignment(horizontal="center", vertical="center")

        # Detect numeric columns from the data body (majority of non-empty cells numeric -> numeric col).
        numeric_col = {}
        for col in range(1, max_col + 1):
            num, total = 0, 0
            for row in range(header_row + 1, max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None or v == "":
                    continue
                total += 1
                if _is_number(v):
                    num += 1
            numeric_col[col] = (total > 0 and num >= total / 2)

        # --- header row ---
        for col in range(1, max_col + 1):
            c = ws.cell(row=header_row, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_mid
            c.border = header_border

        # --- data rows: zebra (recomputed, idempotent), font, alignment, border ---
        for i, row in enumerate(range(header_row + 1, max_row + 1)):
            striped = (i % 2 == 1)  # first data row plain, second tinted, ...
            for col in range(1, max_col + 1):
                c = ws.cell(row=row, column=col)
                c.font = body_font
                c.fill = zebra_fill if striped else no_fill
                c.alignment = right_mid if numeric_col.get(col) else left_mid
                c.border = border

        # --- column widths: fit to content, cap 50 ---
        for col in range(1, max_col + 1):
            longest = 0
            for row in range(header_row, max_row + 1):
                cell = ws.cell(row=row, column=col)
                v = cell.value
                if v is None:
                    continue
                # Count display width: CJK chars are ~2 cells wide in Excel's default font.
                s = str(v)
                width = sum(2 if ord(ch) > 0x2E7F else 1 for ch in s)
                # 把关直修(v1.7 审美关真机撞出):cell 若带数字格式(千分位/货币/小数/百分号,write_excel
                # 落笔时设的),显示宽度按【格式化后】算——否则 beautify 重算列宽时把 write_excel 已算对的
                # 宽度按原始值 "1250" 盖窄,渲染成 ######。与 document.py write_excel 同一近似公式。
                fmt = str(cell.number_format or "")
                if isinstance(v, (int, float)) and fmt and fmt != "General":
                    try:
                        dec = 2 if "0.00" in fmt else (1 if "0.0" in fmt else 0)
                        body = f"{abs(float(v)):,.{dec}f}"
                        symbol = 1 if any(x in fmt for x in ("¥", "$", "％", "%")) else 0
                        sign = 1 if float(v) < 0 else 0
                        width = max(width, len(body) + symbol + sign)
                    except Exception:
                        pass
                longest = max(longest, width)
            # +2 padding; floor 8, cap 50.
            ws.column_dimensions[get_column_letter(col)].width = min(max(longest + 2, 8), 50)

        # --- freeze the header row (everything at/above header stays; scroll starts below) ---
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # --- auto-filter over the used range (header .. last data row) ---
        first = f"A{header_row}"
        last = f"{get_column_letter(max_col)}{max_row}"
        ws.auto_filter.ref = f"{first}:{last}"

        wb.save(path)
        wb.close()
        return {
            "success": True,
            "path": os.path.abspath(path),
            "output_path": os.path.abspath(path),
            "sheet": ws.title,
            "style": resolved_style,
            "rows": max_row,
            "cols": max_col,
        }
    except Exception as e:  # noqa: BLE001
        return {"error": f"美化失败：{e}"}


def _rich_title(text: str, font_name: str):
    """Build an openpyxl chart-axis `Title` whose text is `text` rendered in the CJK `font_name`.

    We hand-build the RichText (not the bare-string shortcut) so the axis-title glyphs also carry the
    latin+ea typeface — otherwise Office draws 中文 axis titles in its default Latin face while the
    surrounding data is 微软雅黑. Returns a Title object, or None if openpyxl's chart-text classes are
    unavailable (defensive — axis titles then simply degrade to unstyled)."""
    try:
        from openpyxl.chart.title import Title
        from openpyxl.chart.text import RichText, Text
        from openpyxl.drawing.text import (
            Paragraph, ParagraphProperties, CharacterProperties, Font as DrawFont, RegularTextRun,
        )
    except Exception:  # noqa: BLE001
        return None
    cp = CharacterProperties(latin=DrawFont(typeface=font_name), ea=DrawFont(typeface=font_name))
    run = RegularTextRun(rPr=cp, t=str(text))
    para = Paragraph(pPr=ParagraphProperties(defRPr=cp), r=[run])
    # Title.tx is a <c:tx> wrapper (openpyxl.chart.text.Text) that HOLDS the RichText — passing the
    # RichText straight in raises a type error, so wrap it.
    return Title(tx=Text(rich=RichText(p=[para])))


def _apply_chart_font(chart, font_name: str):
    """Set the CJK typeface on the chart title / both axes / legend so text is 微软雅黑, not the Office
    default Latin face. Uses openpyxl's txPr/RichText with a latin+ea CharacterProperties. Best-effort:
    any missing sub-element is silently skipped (openpyxl chart objects vary by type)."""
    try:
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import (
            Paragraph, ParagraphProperties, CharacterProperties, Font as DrawFont,
        )
    except Exception:  # noqa: BLE001
        return

    def _txpr():
        cp = CharacterProperties(latin=DrawFont(typeface=font_name),
                                 ea=DrawFont(typeface=font_name))
        return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    # title
    try:
        if chart.title is not None and getattr(chart.title, "tx", None) is not None \
                and getattr(chart.title.tx, "rich", None) is not None:
            rich = chart.title.tx.rich
            cp = CharacterProperties(latin=DrawFont(typeface=font_name),
                                     ea=DrawFont(typeface=font_name))
            for para in rich.p:
                if para.pPr is None:
                    para.pPr = ParagraphProperties()
                para.pPr.defRPr = cp
    except Exception:  # noqa: BLE001
        pass
    # axes (tick-label / axis-title font). We set txPr on the axis itself so the SCALE numbers and
    # category labels also render in 微软雅黑; the axis *title* text font is handled by _rich_title.
    for axis_attr in ("x_axis", "y_axis"):
        try:
            axis = getattr(chart, axis_attr, None)
            if axis is not None:
                axis.txPr = _txpr()
        except Exception:  # noqa: BLE001
            pass
    # legend
    try:
        if getattr(chart, "legend", None) is not None:
            chart.legend.txPr = _txpr()
    except Exception:  # noqa: BLE001
        pass


def _first_col_header_text(ws, min_col, min_row):
    """The header cell text of the category (first) column — used to auto-derive x_title.
    Returns a str (possibly empty) — never raises."""
    try:
        v = ws.cell(row=min_row, column=min_col).value
        return "" if v is None else str(v).strip()
    except Exception:  # noqa: BLE001
        return ""


def _series_header_texts(ws, min_col, max_col, min_row):
    """The header cell texts of the series (non-category) columns — used to auto-derive y_title when a
    single series carries it. Returns a list[str] (skips blanks) — never raises."""
    out = []
    try:
        for col in range(min_col + 1, max_col + 1):
            v = ws.cell(row=min_row, column=col).value
            if v is not None and str(v).strip():
                out.append(str(v).strip())
    except Exception:  # noqa: BLE001
        pass
    return out


def _parse_range(data_range: str):
    """Validate and split an 'A1:B10' range. Return (info_dict, None) or (None, error_dict).

    info: {min_col, min_row, max_col, max_row}. Rejects single-cell / malformed / reversed ranges.
    """
    try:
        from openpyxl.utils.cell import range_boundaries
    except Exception as e:  # noqa: BLE001
        return None, {"error": f"无法解析区域（openpyxl 缺失）：{e}"}
    if not isinstance(data_range, str) or ":" not in data_range:
        return None, {"error": f"data_range 非法：{data_range!r}，应形如 'A1:B10'"}
    try:
        min_col, min_row, max_col, max_row = range_boundaries(data_range.strip())
    except Exception as e:  # noqa: BLE001
        return None, {"error": f"data_range 非法：{data_range!r}（{e}），应形如 'A1:B10'"}
    if None in (min_col, min_row, max_col, max_row):
        return None, {"error": f"data_range 非法：{data_range!r}，需为完整的矩形区域，如 'A1:B10'"}
    if max_row <= min_row:
        return None, {"error": f"data_range {data_range!r} 至少要有表头行 + 一行数据"}
    return {"min_col": min_col, "min_row": min_row, "max_col": max_col, "max_row": max_row}, None


@mcp.tool(audit=True)
def excel_chart(
    path: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    title: str,
    target_cell: str = "H2",
    x_title: str | None = None,
    y_title: str | None = None,
    allow_protected: bool = False,
) -> dict:
    """Insert a native chart (bar | line | pie) into an existing .xlsx, coloured from design tokens.

    The first row of data_range is treated as series names (headers) and the first column as the
    category axis (labels). Chart series colours come from the chosen style's chart_palette.

    v1.7.1 — 坐标轴标题 (用户反馈「Excel 图表好多都没有 X/Y 轴单位」):
      * x_title / y_title 显式设置横 / 纵轴标题 (如「季度」「销售额(万元)」)。
      * 缺省自动推导 (仅在该参数为 None 时):
          - x_title  ← data_range 首列 (类别列) 的表头单元格文本；
          - y_title  ← 单系列时取该系列的表头文本；多系列时留空 (由图例承担轴含义)。
        传空字符串 '' 可显式关闭某轴标题的自动推导。
      * 轴标题与刻度文字字体走令牌 body_font (微软雅黑)，与图内其它文字一致。
      * 常见坑规避: 设 axis.title 后显式 `x_axis.delete = False` / `y_axis.delete = False`，否则
        openpyxl 默认可能把轴 (连同标题) 隐藏。
      * 饼图无坐标轴，x_title / y_title 被忽略。

    Args:
        path: Path to an existing .xlsx.
        sheet: Worksheet name the data lives on (the chart is placed on the same sheet).
        chart_type: 'bar' | 'line' | 'pie'.
        data_range: Data area as 'A1:B10' — first row = series header, first column = categories.
        title: Chart title (中文 OK).
        target_cell: Anchor cell for the chart's top-left corner (default 'H2').
        x_title: 横轴标题。None = 自动推导 (首列表头)；'' = 不加横轴标题。饼图忽略。
        y_title: 纵轴标题。None = 自动推导 (单系列取系列表头，多系列留空)；'' = 不加纵轴标题。饼图忽略。
        allow_protected: Override the protected-system-root guard on the destination (default off).

    Returns:
        dict with 'success', 'path', 'output_path', 'sheet', 'chart_type', 'anchor', 'x_title',
        'y_title' (the axis titles actually applied — '' when none). On failure a {'error': <中文人话>}
        dict (bad type / bad range / missing sheet / import failure).
    """
    if not str(path).lower().endswith(".xlsx"):
        return {"error": "path 必须以 .xlsx 结尾"}
    if not os.path.exists(path):
        return {"error": f"文件不存在：{path}"}
    ctype = str(chart_type).strip().lower()
    if ctype not in ("bar", "line", "pie"):
        return {"error": f"chart_type 非法：{chart_type!r}，仅支持 bar | line | pie"}
    guard = _protected_write_guard(path, allow_protected)
    if guard:
        return guard

    mods, err = _require_openpyxl()
    if err:
        return err

    rng, rerr = _parse_range(data_range)
    if rerr:
        return rerr

    try:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint
    except Exception as e:  # noqa: BLE001
        return {"error": f"图表功能需要 openpyxl.chart（应随离线包安装）。导入失败：{e}"}

    # style is not a param of this tool per the contract; charts use the default (business) palette + font.
    biz = style_tokens.get_style("business")
    palette = biz["chart_palette"]
    chart_font_name = biz["body_font"]

    try:
        wb = mods["load_workbook"](path)
        if sheet not in wb.sheetnames:
            wb.close()
            return {"error": f"工作表不存在：{sheet}。可用：{wb.sheetnames}"}
        ws = wb[sheet]

        min_col, min_row = rng["min_col"], rng["min_row"]
        max_col, max_row = rng["max_col"], rng["max_row"]

        if ctype == "bar":
            chart = BarChart()
            chart.type = "col"
        elif ctype == "line":
            chart = LineChart()
        else:
            chart = PieChart()

        chart.title = str(title)
        chart.style = 2
        # v1.7: align chart title / axis / legend fonts to the token body font (微软雅黑) so the chart
        # doesn't render its text in Office's default Latin face while data around it is 微软雅黑.
        _apply_chart_font(chart, chart_font_name)

        # Categories = first column (excluding the header cell). Data = remaining columns incl. header
        # row so series pick up their names.
        cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=max_col, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # --- v1.7.1 坐标轴标题 (饼图无轴，跳过) ---
        applied_x, applied_y = "", ""
        if ctype != "pie":
            # x_title: explicit unless None → auto-derive from the category-column header.
            if x_title is None:
                applied_x = _first_col_header_text(ws, min_col, min_row)
            else:
                applied_x = str(x_title)
            # y_title: explicit unless None → auto-derive from the single series header (blank if
            # multi-series, where the legend already names each series).
            if y_title is None:
                headers = _series_header_texts(ws, min_col, max_col, min_row)
                applied_y = headers[0] if len(headers) == 1 else ""
            else:
                applied_y = str(y_title)

            # delTitle 陷阱: 设 axis.title 后 axis.delete 若为 True 轴会连标题一起隐藏；显式关掉。
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            if applied_x:
                t = _rich_title(applied_x, chart_font_name)
                chart.x_axis.title = t if t is not None else applied_x
            if applied_y:
                t = _rich_title(applied_y, chart_font_name)
                chart.y_axis.title = t if t is not None else applied_y

        # Apply palette colours.
        if ctype == "pie":
            # One series, colour each slice (data point) from the palette.
            if chart.series:
                s = chart.series[0]
                n_points = max_row - min_row  # number of categories
                for idx in range(n_points):
                    dp = DataPoint(idx=idx)
                    dp.graphicalProperties.solidFill = palette[idx % len(palette)]
                    s.data_points.append(dp)
        else:
            for i, s in enumerate(chart.series):
                color = palette[i % len(palette)]
                s.graphicalProperties.solidFill = color
                if ctype == "line":
                    s.graphicalProperties.line.solidFill = color
                    s.smooth = False

        chart.height = 8   # cm
        chart.width = 15   # cm

        try:
            ws.add_chart(chart, str(target_cell))
        except Exception as e:  # noqa: BLE001
            wb.close()
            return {"error": f"target_cell 非法：{target_cell!r}（{e}），应形如 'H2'"}

        wb.save(path)
        wb.close()
        return {
            "success": True,
            "path": os.path.abspath(path),
            "output_path": os.path.abspath(path),
            "sheet": sheet,
            "chart_type": ctype,
            "anchor": str(target_cell),
            "x_title": applied_x,
            "y_title": applied_y,
        }
    except Exception as e:  # noqa: BLE001
        return {"error": f"插入图表失败：{e}"}
